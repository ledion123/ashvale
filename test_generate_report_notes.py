"""Self-check: generate_report()'s Notes column uses the same item-level LOLER/PUWER
serial parsing (bare numbers, implied prefix, photo-uploaded detection) as the
Dashboard and /api/loler-check, instead of the old blind extract_serials(all_text).
Run: python test_generate_report_notes.py"""
import io
import os

os.environ.setdefault("SAFETYCULTURE_API_TOKEN", "dummy")

import openpyxl
import app as appmod

loler_tid = next(tid for tid, tkey in appmod.TEMPLATES.items() if tkey == "LOLER")
puwer_tid = next(tid for tid, tkey in appmod.TEMPLATES.items() if tkey == "PUWER_REGISTER")


def fake_search_audits(template_id, modified_after, modified_before):
    if template_id == loler_tid:
        return [{"audit_id": "loler-rich"}, {"audit_id": "loler-photo"}]
    if template_id == puwer_tid:
        return [{"audit_id": "puwer-bare"}]
    return []


def fake_fetch_audits_parallel(audit_ids, max_workers=50):
    details = {
        "loler-rich": {
            "audit_data": {"site": {"name": "AB20 Rich Site"}, "date_completed": "2026-07-29T10:00:00Z"},
            "items": [{"label": "Machine Forks", "responses": {"text": "16,32"}}],
        },
        "loler-photo": {
            "audit_data": {"site": {"name": "BG223 Photo Site"}, "date_completed": "2026-07-29T10:00:00Z"},
            "items": [{"label": "4 leg chain", "responses": {}, "media": [{"media_id": "x"}]}],
        },
        "puwer-bare": {
            "audit_data": {"site": {"name": "DAN004 Bare Site"}, "date_completed": "2026-07-29T10:00:00Z"},
            "items": [{"label": "0.75 Ton Excavator", "responses": {"text": "802"}}],
        },
    }
    return {aid: details[aid] for aid in audit_ids if aid in details}


appmod.sc.search_audits = fake_search_audits
appmod.sc.fetch_audits_parallel = fake_fetch_audits_parallel

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet2"
ws["A1"] = "SITE"
ws["B1"] = "Job"
rows = {
    "AB20 Rich Site": ("AB20 Rich Site", "AB20"),
    "BG223 Photo Site": ("BG223 Photo Site", "BG223"),
    "DAN004 Bare Site": ("DAN004 Bare Site", "DAN004"),
}
for i, (site, job) in enumerate(rows.values(), start=2):
    ws.cell(row=i, column=1, value=site)
    ws.cell(row=i, column=2, value=job)

buf = io.BytesIO()
wb.save(buf)
excel_bytes = buf.getvalue()

out = appmod.generate_report("2026-07-27", "2026-08-02", excel_bytes, pdf_bytes=None)
out_wb = openpyxl.load_workbook(out)
out_ws = out_wb["Sheet2"]


def note_for(row_num):
    return out_ws.cell(row=row_num, column=appmod.NOTES_COL).value or ""


# Row 2: bare-number LOLER text ("16,32" under "Machine Forks") is now parsed via
# extract_loler_item_serials -> found_serials non-empty -> no false "No serial no." note.
assert "No serial no." not in note_for(2), note_for(2)

# Row 3: LOLER item has media but no text -> report_uploaded, not a false "missing everything".
assert "Report uploaded as a file" in note_for(3), note_for(3)

# Row 4: bare "802" under a PUWER "0.75 Ton Excavator" item is now parsed via
# extract_puwer_item_serials (implied ASH prefix) -> no false "No serial no." note.
assert "No serial no." not in note_for(4), note_for(4)

print("OK: generate_report() Notes column reflects item-level LOLER/PUWER parsing and photo-uploaded detection")
