import hmac
import io
import json as _json
import os
import re
import tempfile
import requests
import pdfplumber
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta, timezone
from flask import Flask, request, send_file, jsonify, send_from_directory, Response
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from dotenv import load_dotenv
load_dotenv()

import sc_client as sc

app = Flask(__name__, static_folder="static", static_url_path="/static")
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024  # 20MB — plenty for a weekly Excel/PDF upload

APP_USER = os.environ.get("APP_USER")
APP_PASS = os.environ.get("APP_PASS")
if not APP_USER or not APP_PASS:
    app.logger.warning("APP_USER/APP_PASS not set - Basic Auth is DISABLED, app is open to anyone")


@app.before_request
def require_auth():
    """Gate the whole app behind HTTP Basic Auth if APP_USER/APP_PASS are set.
    No-op when unset, so local dev without those env vars keeps working."""
    if not APP_USER or not APP_PASS:
        return
    auth = request.authorization
    valid = auth and hmac.compare_digest(auth.username, APP_USER) and hmac.compare_digest(auth.password, APP_PASS)
    if not valid:
        return Response("Authentication required", 401, {"WWW-Authenticate": 'Basic realm="Ashvale"'})

TEMPLATES = {
    "template_7dbbb416041a44459216a2a0ba02bb10": "LOLER",
    "template_9de384b691994b498011b166402163d3": "TELEHAND",
    "template_ce94e9eade8f4a7aaa3c894cdf4b3934": "ROLLER",
    "template_785b009c0581405a86b870ecd52cfa79": "EXCAVATOR",
    "template_4d4487f16b27468ab069be1262b05376": "DUMPER",
    # Large sites use daily check sheets bundled into one weekly upload instead of the
    # normal template above — either one satisfies the same EXCAVATOR/DUMPER column.
    "template_5c78491fe04347adbcbeea8b39828937": "EXCAVATOR",
    "template_ce52353281b64255965183cd769116bd": "DUMPER",
    "template_0a8a57e828e746f782b2659da47f398d": "PUWER_REGISTER",
    "template_dbd9e89576dd43898fdd2f89e091dead": "SITE SUP",
    "template_6d565926a30944f2927a7760c343f4ee": "HAVS",
    "template_277834af991b482c945030f0f936f5a9": "TOOLBOX",
}

# PUWER_REGISTER completion marks all five machine-type columns
PUWER_COLUMNS = ("PUWER", "EXCAVATOR", "DUMPER", "ROLLER", "TELEHAND")

# Columns that have their own individual inspection template, separate from PUWER_REGISTER
INDIVIDUAL_COVERABLE = {"EXCAVATOR", "DUMPER", "ROLLER", "TELEHAND"}

DISPLAY_COLUMNS = ["EXCAVATOR", "LOLER", "DUMPER", "ROLLER", "TELEHAND", "PUWER", "SITE SUP", "HAVS", "TOOLBOX"]

# 0-based column indexes in Sheet2
COL_COMPLIANCE = {
    "PUWER":     2,
    "LOLER":     3,
    "SITE SUP":  4,
    "EXCAVATOR": 5,
    "DUMPER":    6,
    "ROLLER":    7,
    "TELEHAND":  8,
    "HAVS":      9,
    "TOOLBOX":   10,
}
NOTES_COL = 14

PREFIX_TYPE = {
    "ROR": "ROLLER",
    "TL":  "TELEHAND",
    "SL":  "LOLER",
    "SH":  "LOLER",
    "FK":  "LOLER",
    "BG":  "LOLER",
    "TS":  "LOLER",
}
ASH_KEYWORDS = {
    "EXCAVATOR":   "EXCAVATOR",
    "DIGGER":      "EXCAVATOR",
    "JCB":         "EXCAVATOR",
    "DUMPER":      "DUMPER",
    "ROLLER":      "ROLLER",
    "TELEHANDLER": "TELEHAND",
    "TELEHAND":    "TELEHAND",
}

SERIAL_TOKEN_RE = re.compile(r'\b(ASH|ROR|TL|SL|SH|FK|BG|TS)\s*([\dA-Z][\dA-Z/]*)', re.IGNORECASE)
JOB_CODE_RE     = re.compile(r'\[([A-Z]{2,4}\d{2,})\]')

GREEN  = PatternFill("solid", fgColor="27AE60")
RED    = PatternFill("solid", fgColor="C0392B")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
W_FONT = Font(bold=True, color="FFFFFF", size=10)


# ── Helpers ───────────────────────────────────────────────

def normalize_job(code):
    code = str(code).strip().upper()
    m = re.match(r'^([A-Z]+)(\d+)$', code)
    return (m.group(1) + str(int(m.group(2)))) if m else code

def norm_name(s):
    s = re.sub(r'[^a-z0-9 ]', ' ', s.lower())
    return ' '.join(s.split())

def extract_serials(text):
    serials = set()
    for m in SERIAL_TOKEN_RE.finditer(text):
        prefix = m.group(1).upper()
        for part in m.group(2).split('/'):
            part = part.strip().rstrip('(').strip()
            if part and any(c.isdigit() for c in part):
                serials.add(prefix + part.upper())
    return serials

# Both PUWER Register and LOLER list serials in each item's own response text,
# often as bare numbers with no prefix (e.g. "802\n61\n43" under a PUWER Excavator
# item, or "16,32,40" under a LOLER "Machine Forks" item) since the item's own
# question already implies the equipment category.
def _implied_prefix_serials(item_text, implied_prefix):
    """
    Parse serials an inspector typed into an item's response text. Explicitly-
    prefixed tokens (e.g. "ROR 17", "SL8717") are parsed via the existing
    extract_serials(). Bare numeric tokens (e.g. "802", "19") fall back to the
    item's implied prefix — but only when the WHOLE token is numeric, so free-text
    noise like "Hire" or "Not in use" is safely skipped rather than guessed at.
    """
    if not item_text:
        return set()
    serials = set()
    for token in re.split(r'[,/\n]', item_text):
        token = token.strip()
        if not token:
            continue
        explicit = extract_serials(token)
        if explicit:
            serials.update(explicit)
        elif implied_prefix and token.isdigit():
            serials.add(implied_prefix + token)
    return serials

PUWER_IMPLIED_PREFIX = {
    "EXCAVATOR": "ASH",
    "DUMPER": "ASH",
    "ROLLER": "ROR",
    "TELEHAND": "TL",
}

def extract_puwer_item_serials(mtype, item_text):
    """PUWER Register item text -> serials, using the item's machine-type category
    (Excavator/Dumper -> ASH, Roller -> ROR, Telehandler -> TL) as the implied prefix."""
    return _implied_prefix_serials(item_text, PUWER_IMPLIED_PREFIX.get(mtype))

# LOLER item labels -> implied prefix, keyed by keyword found in the item's own
# label. Concrete Bucket/Skip, Lifting Straps, and Harness use codes the app has
# never recognised for anything and are deliberately left unmapped rather than
# guessed at (same as the Pedestrian/Trench Roller gap noted elsewhere).
LOLER_ITEM_KEYWORDS = {
    "CHAIN": "SL",
    "SHACKLE": "SH",
    "FORK": "FK",
    "BLOCK GRAB": "BG",
    "TIPPING SKIP": "TS",
}

def extract_loler_item_serials(item_label, item_text):
    """LOLER item text -> serials, using a keyword match on the item's own label
    (chain -> SL, shackle -> SH, fork -> FK, block grab -> BG, tipping skip -> TS)
    as the implied prefix for bare numeric entries."""
    label_upper = (item_label or "").upper()
    prefix = next((p for kw, p in LOLER_ITEM_KEYWORDS.items() if kw in label_upper), None)
    return _implied_prefix_serials(item_text, prefix)

def extract_machine_id(audit_name):
    """
    Individual EXCAVATOR/DUMPER/ROLLER audit titles follow "DATE / SERIAL / SITE NAME"
    (e.g. "20 Jul 2026 / ASH 1419 / BG218 The Queens Drive, Mill End"). Only the middle
    segment is scanned for a serial — never the site-name segment, since some job codes
    (e.g. "BG218") collide with recognised serial prefixes (BG -> LOLER) and would
    otherwise be misread as a machine serial. Returns None if the title isn't in this
    3-part shape (e.g. TELEHAND/LOLER titles are just "DATE / SITE NAME") or no serial
    is found in the middle segment.
    """
    parts = audit_name.split(" / ")
    if len(parts) != 3:
        return None
    serials = extract_serials(parts[1])
    return next(iter(serials), None)

def current_week_range():
    today = datetime.now(timezone.utc).date()
    monday = today - timedelta(days=today.weekday())
    return str(monday), str(today)

def to_iso(date_str, end=False):
    t = "T23:59:59Z" if end else "T00:00:00Z"
    return f"{date_str}{t}"

def has_file(f):
    """True only if a file was actually selected, not just an empty form field."""
    return bool(f and f.filename)

def valid_date_format(*dates):
    try:
        for d in dates:
            datetime.strptime(d, "%Y-%m-%d")
        return True
    except ValueError:
        return False

def search_templates(templates, modified_after, modified_before, max_workers=9):
    """
    Search all given templates ({template_id: tkey}) in parallel.
    Returns (audit_to_tkey, errors) — errors is a list of {"template", "error"}
    for templates whose search failed. Raises RuntimeError if EVERY template
    failed (e.g. a bad/expired SAFETYCULTURE_API_TOKEN), so callers surface a
    real error instead of silently reporting zero results as if compliant.
    """
    audit_to_tkey = {}
    errors = []

    def _search(tid_tkey):
        tid, tkey = tid_tkey
        try:
            return tkey, sc.search_audits(tid, modified_after, modified_before), None
        except Exception as e:
            return tkey, [], str(e)

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        for tkey, audits, error in ex.map(_search, templates.items()):
            if error:
                errors.append({"template": tkey, "error": error})
            for a in audits:
                audit_to_tkey[a["audit_id"]] = tkey

    if templates and len(errors) == len(templates):
        raise RuntimeError(f"SafetyCulture search failed for all templates: {errors[0]['error']}")

    return audit_to_tkey, errors

def group_site_key(sc_name):
    """Derive (job_code_or_name, display_name, job_code) from a SafetyCulture site name."""
    first = sc_name.split()[0] if sc_name else ""
    if re.match(r'^[A-Z]{1,4}\d{2,}$', first, re.I):
        return normalize_job(first), sc_name, normalize_job(first)
    return sc_name, sc_name, ""


# ── PDF parsing ───────────────────────────────────────────

def parse_pdf(pdf_bytes):
    site_machines = {}
    name_to_job   = {}

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue
                jm = JOB_CODE_RE.search(line)
                if not jm:
                    continue
                raw_job   = jm.group(1).upper()
                norm_job  = normalize_job(raw_job)
                site_name = line[:jm.start()].strip()
                if norm_job not in site_machines:
                    site_machines[norm_job] = {}
                    if site_name:
                        name_to_job[norm_name(site_name)] = norm_job
                line_upper = line.upper()
                serials    = extract_serials(line)
                for serial in serials:
                    prefix = re.match(r'^([A-Z]+)', serial)
                    if not prefix:
                        continue
                    pfx = prefix.group(1)
                    if pfx in PREFIX_TYPE:
                        mtype = PREFIX_TYPE[pfx]
                    elif pfx == "ASH":
                        mtype = next((mt for kw, mt in ASH_KEYWORDS.items() if kw in line_upper), "EXCAVATOR")
                    else:
                        continue
                    site_machines[norm_job].setdefault(mtype, set()).add(serial)

    return site_machines, name_to_job


# ── Excel template helpers ────────────────────────────────

def build_lookup(ws):
    job_lookup  = {}
    name_lookup = {}
    header_row  = 1
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=2):
        if row[0].value and str(row[0].value).strip().upper() == "SITE":
            header_row = row[0].row
            break
    for row in ws.iter_rows(min_row=header_row + 1, max_col=2):
        site_val = str(row[0].value).strip() if row[0].value else ""
        job_val  = str(row[1].value).strip() if row[1].value else ""
        if not site_val:
            continue
        rn = row[0].row
        if job_val:
            job_lookup[normalize_job(job_val)] = rn
        name_lookup[site_val.lower()] = rn
    return job_lookup, name_lookup, header_row

def find_row(sc_site, job_lookup, name_lookup):
    parts  = sc_site.strip().split()
    sc_job = parts[0].upper() if parts and re.match(r'^[A-Z]{1,4}\d{2,}$', parts[0], re.I) else ""
    if sc_job and normalize_job(sc_job) in job_lookup:
        return job_lookup[normalize_job(sc_job)]
    if sc_site.lower() in name_lookup:
        return name_lookup[sc_site.lower()]
    sc_words = set(sc_site.lower().split()) - {sc_job.lower()}
    best_row, best = None, 0
    for name, rn in name_lookup.items():
        c = len(sc_words & set(name.split()))
        if c > best:
            best, best_row = c, rn
    return best_row if best >= 2 else None

def mark_cell(ws, row_num, col_idx, done=True):
    cell = ws.cell(row=row_num, column=col_idx + 1)
    orig = str(cell.value).strip().upper() if cell.value else ""
    if orig == "N/A":
        cell.value = ""
        cell.fill  = PatternFill()
        return
    cell.value     = "Y" if done else "N"
    cell.fill      = GREEN if done else RED
    cell.font      = W_FONT
    cell.alignment = CENTER

def append_note(notes_dict, row_num, msg):
    notes_dict.setdefault(row_num, [])
    if msg not in notes_dict[row_num]:
        notes_dict[row_num].append(msg)


# ── Excel report generation (unchanged) ──────────────────

def generate_report(from_date, to_date, excel_bytes, pdf_bytes=None):
    modified_after  = datetime.strptime(from_date, "%Y-%m-%d").strftime("%Y-%m-%dT00:00:00Z")
    modified_before = datetime.strptime(to_date,   "%Y-%m-%d").strftime("%Y-%m-%dT23:59:59Z")

    pdf_machines = {}
    pdf_name_map = {}
    if pdf_bytes:
        pdf_machines, pdf_name_map = parse_pdf(pdf_bytes)

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.active
    job_lookup, name_lookup, header_row = build_lookup(ws)
    row_to_normjob = {rn: nj for nj, rn in job_lookup.items()}

    matched   = {}
    row_notes = {}

    audit_to_tkey, _ = search_templates(TEMPLATES, modified_after, modified_before)
    details = sc.fetch_audits_parallel(list(audit_to_tkey.keys()), max_workers=50)

    for audit_id, d in details.items():
        tkey    = audit_to_tkey[audit_id]
        ad      = d.get("audit_data", {})
        items   = d.get("items", [])
        sc_site = ad.get("site", {}).get("name", "").strip()

        row_num = find_row(sc_site, job_lookup, name_lookup)
        if row_num is None:
            continue

        matched.setdefault(row_num, set())
        row_notes.setdefault(row_num, [])
        display = "PUWER" if tkey == "PUWER_REGISTER" else tkey

        if tkey == "PUWER_REGISTER":
            for col_key in PUWER_COLUMNS:
                mark_cell(ws, row_num, COL_COMPLIANCE[col_key], True)
                matched[row_num].add(col_key)
        elif tkey in COL_COMPLIANCE:
            mark_cell(ws, row_num, COL_COMPLIANCE[tkey], True)
            matched[row_num].add(tkey)

        if tkey in ("LOLER", "EXCAVATOR", "DUMPER", "ROLLER", "TELEHAND") or tkey == "PUWER_REGISTER":
            found_serials = set()
            has_media = False
            if tkey == "LOLER":
                for item in items:
                    item_text = (item.get("responses", {}).get("text") or "") + " " + (item.get("note") or "")
                    found_serials |= extract_loler_item_serials(item.get("label"), item_text)
                    if item.get("media"):
                        has_media = True
            elif tkey == "PUWER_REGISTER":
                for item in items:
                    label_upper = (item.get("label") or "").upper()
                    mtype = next((mt for kw, mt in ASH_KEYWORDS.items() if kw in label_upper), None)
                    if not mtype:
                        continue
                    item_text = item.get("responses", {}).get("text") or ""
                    item_serials = extract_puwer_item_serials(mtype, item_text)
                    if item_serials:
                        found_serials |= item_serials
                    elif item.get("media"):
                        has_media = True
            else:
                all_text = ""
                for item in items:
                    responses = item.get("responses", {})
                    all_text += " " + (responses.get("text") or "")
                    all_text += " " + (item.get("note") or "")
                found_serials = extract_serials(all_text)
                # Item text/notes are often left blank; the audit title reliably has the
                # machine serial for individual EXCAVATOR/DUMPER/ROLLER/TELEHAND audits.
                machine_id = extract_machine_id(ad.get("name", ""))
                if machine_id:
                    found_serials.add(machine_id)

            # A whole report attached as a photo instead of the checklist being filled
            # in — nothing to read, don't guess (same treatment as the Dashboard/LOLER-check).
            report_uploaded = has_media and not found_serials

            if pdf_bytes:
                norm_job   = row_to_normjob.get(row_num, "")
                registered = set()
                if norm_job and norm_job in pdf_machines:
                    if tkey == "PUWER_REGISTER":
                        for mt in ("EXCAVATOR", "DUMPER", "ROLLER", "TELEHAND"):
                            registered |= pdf_machines[norm_job].get(mt, set())
                    else:
                        registered = pdf_machines[norm_job].get(tkey, set())
                if report_uploaded:
                    append_note(row_notes, row_num, f"Report uploaded as a file — not auto-verified ({display})")
                elif not found_serials:
                    append_note(row_notes, row_num, f"No serial no. on {display}")
                else:
                    missing = registered - found_serials
                    unknown = found_serials - registered
                    if missing:
                        append_note(row_notes, row_num, f"Not inspected ({display}): {', '.join(sorted(missing))}")
                    if unknown:
                        append_note(row_notes, row_num, f"Not in register ({display}): {', '.join(sorted(unknown))}")
            else:
                if report_uploaded:
                    append_note(row_notes, row_num, f"Report uploaded as a file — not auto-verified ({display})")
                elif not found_serials:
                    append_note(row_notes, row_num, f"No serial no. on {display}")

        for item in items:
            selected = item.get("responses", {}).get("selected", [])
            status   = selected[0].get("label", "").strip() if selected else ""
            if status == "At Risk":
                label = (item.get("label") or "").strip()
                append_note(row_notes, row_num, f"At Risk: {label[:40]} ({display})")

    for row in ws.iter_rows(min_row=header_row + 1):
        if not row[0].value:
            continue
        row_num  = row[0].row
        done     = matched.get(row_num, set())
        norm_job = row_to_normjob.get(row_num, "")
        site_machines_here = pdf_machines.get(norm_job, {}) if pdf_bytes else None

        for col_key, col_idx in COL_COMPLIANCE.items():
            if col_key in done:
                continue
            if site_machines_here is not None and col_key in ("EXCAVATOR", "DUMPER", "ROLLER", "TELEHAND", "LOLER"):
                if col_key not in site_machines_here:
                    cell = ws.cell(row=row_num, column=col_idx + 1)
                    orig = str(cell.value).strip().upper() if cell.value else ""
                    if orig == "N/A":
                        cell.value = ""
                        cell.fill  = PatternFill()
                    continue
            mark_cell(ws, row_num, col_idx, False)

    for row_num, notes in row_notes.items():
        if not notes:
            continue
        cell     = ws.cell(row=row_num, column=NOTES_COL)
        existing = str(cell.value).strip() if cell.value else ""
        new_part = "; ".join(n for n in notes if n not in existing)
        if new_part:
            cell.value = (existing + "; " + new_part).strip("; ") if existing else new_part

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ── LOLER item-level coverage check ──────────────────────

def check_loler_coverage(from_date, to_date, pdf_bytes):
    """
    Cross-reference completed LOLER audits against the plant register PDF:
    for each site, which registered lifting-gear serials were actually
    mentioned in the checklist responses (found_in_check) vs not (missing),
    and which mentioned serials aren't in the register at all (unknown).
    """
    site_machines, _ = parse_pdf(pdf_bytes)

    modified_after  = to_iso(from_date)
    modified_before = to_iso(to_date, end=True)

    loler_tid = next(tid for tid, tkey in TEMPLATES.items() if tkey == "LOLER")
    audits  = sc.search_audits(loler_tid, modified_after, modified_before)
    details = sc.fetch_audits_parallel([a["audit_id"] for a in audits], max_workers=20)

    seen_jobs = set()
    results   = []

    for detail in details.values():
        ad        = detail.get("audit_data", {})
        site_name = ad.get("site", {}).get("name", "").strip()
        if not site_name:
            continue

        _, display, job = group_site_key(site_name)
        registered = site_machines.get(job, {}).get("LOLER", set())
        seen_jobs.add(job)

        found = set()
        has_media = False
        for item in detail.get("items", []):
            item_text = (item.get("responses", {}).get("text") or "") + " " + (item.get("note") or "")
            item_serials = extract_loler_item_serials(item.get("label"), item_text)
            if item_serials:
                found.update(item_serials)
            if item.get("media"):
                has_media = True
        # A whole "Loler report" attached as a file (often to the generic "Any
        # Other comments" item, not the equipment items themselves) instead of the
        # per-item checklist being filled in — nothing to read, don't guess.
        report_uploaded = has_media and not found
        missing = registered - found
        unknown = found - registered

        notes = []
        if report_uploaded:
            notes.append("Report uploaded as a file — not auto-verified")
        elif missing:
            notes.append(f"Not inspected: {', '.join(sorted(missing))}")
        if unknown:
            notes.append(f"Not in register: {', '.join(sorted(unknown))}")
        if registered and not found and not report_uploaded:
            notes.append("No serial no. on LOLER")

        authorship = ad.get("authorship", {})
        results.append({
            "site": display,
            "job_code": job,
            "status": "done" if ad.get("date_completed") else "missing",
            "date_completed": ad.get("date_completed"),
            "inspector": authorship.get("author", "") if isinstance(authorship, dict) else "",
            "registered": sorted(registered),
            "found_in_check": sorted(found),
            "missing": sorted(missing) if not report_uploaded else [],
            "unknown": sorted(unknown),
            "report_uploaded": report_uploaded,
            "notes": "; ".join(notes) if notes else "All registered LOLER equipment checked",
        })

    # Sites with LOLER gear registered on the PDF but no LOLER audit at all this week
    for job, machines in site_machines.items():
        registered = machines.get("LOLER", set())
        if registered and job not in seen_jobs:
            results.append({
                "site": job,
                "job_code": job,
                "status": "missing",
                "date_completed": None,
                "inspector": None,
                "registered": sorted(registered),
                "found_in_check": [],
                "missing": sorted(registered),
                "unknown": [],
                "notes": "No LOLER inspection found this week",
            })

    return results


# ── Dashboard cache (file-based, 1-hour TTL, best-effort) ─
# NOTE: this is a per-instance cache, not a shared/distributed one. On
# serverless platforms (e.g. Vercel) each function instance has its own
# ephemeral /tmp, so a cache hit is a same-instance speedup only, not a
# guarantee — concurrent requests may land on different instances and each
# do their own SafetyCulture fetch. That's an acceptable tradeoff (still
# saves real requests when it hits) without adding external cache infra.

_CACHE_FILE = os.path.join(tempfile.gettempdir(), "ashvale_dashboard_cache.json")
_CACHE_TTL  = 3600  # seconds

def _read_cache(cache_key):
    try:
        with open(_CACHE_FILE) as f:
            store = _json.load(f)
        entry = store.get(cache_key)
        if entry and (datetime.now(timezone.utc).timestamp() - entry["ts"]) < _CACHE_TTL:
            return entry["data"]
    except Exception:
        pass
    return None

def _write_cache(cache_key, data):
    try:
        try:
            with open(_CACHE_FILE) as f:
                store = _json.load(f)
        except Exception:
            store = {}
        store[cache_key] = {"ts": datetime.now(timezone.utc).timestamp(), "data": data}
        # Write to a temp file then rename atomically, so a concurrent reader/writer
        # on this same instance never sees a partially-written or lost-update file.
        fd, tmp_path = tempfile.mkstemp(dir=os.path.dirname(_CACHE_FILE), suffix=".tmp")
        try:
            with os.fdopen(fd, "w") as f:
                _json.dump(store, f)
            os.replace(tmp_path, _CACHE_FILE)
        except Exception:
            os.unlink(tmp_path)
            raise
    except Exception:
        pass


# ── Dashboard data builder ────────────────────────────────

def build_dashboard_data(from_date, to_date, force_refresh=False):
    """
    Fetch completed inspections for the selected week.
    Results are cached for 1 hour (best-effort, per-instance — see cache
    section above) to keep subsequent loads on the same instance instant.
    """
    cache_key = f"{from_date}:{to_date}"
    if not force_refresh:
        cached = _read_cache(cache_key)
        if cached:
            return cached

    modified_after  = to_iso(from_date)
    modified_before = to_iso(to_date, end=True)

    week_start = datetime.strptime(from_date, "%Y-%m-%d").date()

    # Search all templates in parallel, then fetch audit details in parallel
    audit_to_tkey, search_errors = search_templates(TEMPLATES, modified_after, modified_before)

    # Fetch all audit details in parallel (50 workers for speed)
    details = sc.fetch_audits_parallel(list(audit_to_tkey.keys()), max_workers=50)

    SERIAL_CHECK_TKEYS = {"LOLER", "EXCAVATOR", "DUMPER", "ROLLER", "TELEHAND", "PUWER_REGISTER"}

    # group_key -> col_key -> {status, last_completed, audit_id, inspector}
    # group_key is normalised job code when available (e.g. "AC23"), else full site name
    site_data    = {}
    site_serials = {}
    # group_key -> set of col_keys satisfied by a non-PUWER_REGISTER (individual) audit
    direct_audit_cols     = {}
    # group_key -> set of col_keys PUWER_REGISTER marked present (answer given, not N/A)
    puwer_register_present = {}
    # group_key -> col_key -> set of distinct machine ids (serial from title, or audit_id
    # fallback) inspected this week via the individual template — not just "was one done"
    machines_checked = {}
    # group_key -> col_key -> set of serials PUWER Register's own item text listed
    puwer_serials = {}
    # group_key -> set of col_keys where PUWER Register had a photo attached instead
    # of a typed serial list — can't be auto-verified, so the cross-check is suppressed
    puwer_photo_only = {}
    # group_keys where LOLER had a whole report attached as a file instead of the
    # per-item checklist being filled in — can't be auto-verified either
    loler_report_uploaded = set()
    # group_key -> col_key -> [{audit_id, machine_id, date_completed, inspector}, ...]
    # every individual audit this week, not just the most-recently-completed one kept
    # in site_data — a site can have many machines of the same type inspected in a week
    individual_audits = {}

    for audit_id, detail in details.items():
        tkey = audit_to_tkey[audit_id]
        ad   = detail.get("audit_data", {})
        site_name = ad.get("site", {}).get("name", "").strip()
        if not site_name:
            continue

        gkey, display, job = group_site_key(site_name)

        date_completed = ad.get("date_completed") or ad.get("date_modified", "")
        authorship = ad.get("authorship", {})
        inspector = authorship.get("author", "") if isinstance(authorship, dict) else ""

        # Determine which display columns this inspection covers
        if tkey == "PUWER_REGISTER":
            col_keys = list(PUWER_COLUMNS)
        else:
            col_keys = [tkey]

        if gkey not in site_data:
            site_data[gkey] = {"_display_name": display, "_job_code": job}

        # PUWER Register answers each machine-type item "N/A" when the site has none
        # of that equipment — figure that out *before* deciding coverage below, so an
        # N/A category doesn't get marked "done" (that's what caused a site with no
        # telehandler to still show a green tick sourced from a register saying so).
        na_categories = set()
        if tkey == "PUWER_REGISTER":
            present = set()
            for item in detail.get("items", []):
                label_upper = (item.get("label") or "").upper()
                mtype = next((mt for kw, mt in ASH_KEYWORDS.items() if kw in label_upper), None)
                if not mtype:
                    continue
                selected = item.get("responses", {}).get("selected", [])
                answer = selected[0].get("label", "").strip().upper() if selected else ""
                if answer == "N/A":
                    na_categories.add(mtype)
                elif answer:
                    present.add(mtype)
                item_text = item.get("responses", {}).get("text") or ""
                item_serials = extract_puwer_item_serials(mtype, item_text)
                if item_serials:
                    puwer_serials.setdefault(gkey, {}).setdefault(mtype, set()).update(item_serials)
                elif item.get("media"):
                    # A photo was attached instead of typing the serial list — we can't
                    # read it (no OCR), so flag it rather than silently treating this
                    # category as if PUWER Register had nothing to say about it.
                    puwer_photo_only.setdefault(gkey, set()).add(mtype)
            puwer_register_present.setdefault(gkey, set()).update(present)

        for col_key in col_keys:
            existing = site_data[gkey].get(col_key)
            # Keep only the most recent inspection per site/column — an N/A verdict
            # is a real answer like any other, so the same freshest-wins rule decides
            # whether it or an older/newer non-N/A entry is the one that's shown.
            if not existing or (date_completed and date_completed > existing["last_completed"]):
                site_data[gkey][col_key] = {
                    "last_completed": date_completed,
                    "audit_id": audit_id,
                    "inspector": inspector,
                    "na": col_key in na_categories,
                }

        if tkey != "PUWER_REGISTER" and tkey in INDIVIDUAL_COVERABLE:
            direct_audit_cols.setdefault(gkey, set()).add(tkey)
            machine_id = extract_machine_id(ad.get("name", "")) or audit_id
            machines_checked.setdefault(gkey, {}).setdefault(tkey, set()).add(machine_id)
            individual_audits.setdefault(gkey, {}).setdefault(tkey, []).append({
                "audit_id": audit_id, "machine_id": machine_id,
                "date_completed": date_completed, "inspector": inspector,
            })

        # LOLER lists serials per-item, often as bare numbers implied by the item's
        # own question (chain/shackle/fork/etc.) — same treatment as PUWER Register.
        if tkey == "LOLER":
            loler_found = set()
            has_media = False
            for item in detail.get("items", []):
                item_text = (item.get("responses", {}).get("text") or "") + " " + (item.get("note") or "")
                item_serials = extract_loler_item_serials(item.get("label"), item_text)
                if item_serials:
                    loler_found.update(item_serials)
                if item.get("media"):
                    has_media = True
            site_serials.setdefault(gkey, {}).setdefault("LOLER", set()).update(loler_found)
            if has_media and not loler_found:
                loler_report_uploaded.add(gkey)
        # Extract serial numbers from all item text/notes for machine-type inspections
        elif tkey in SERIAL_CHECK_TKEYS:
            all_text = " ".join(
                (item.get("responses", {}).get("text") or "") + " " + (item.get("note") or "")
                for item in detail.get("items", [])
            )
            serials = extract_serials(all_text)
            site_serials.setdefault(gkey, {})
            # PUWER_REGISTER covers all machine types — store under "PUWER" only.
            # Individual machine audits (EXCAVATOR, DUMPER, etc.) store under their own key.
            serial_col = "PUWER" if tkey == "PUWER_REGISTER" else tkey
            site_serials[gkey].setdefault(serial_col, set())
            site_serials[gkey][serial_col].update(serials)

    # Build output
    sites_list = []
    for gkey, cols in sorted(site_data.items()):
        site_name = cols["_display_name"]
        job_code  = cols["_job_code"]

        templates_status = {}
        for col_key in DISPLAY_COLUMNS:
            info = cols.get(col_key)

            # Cross-check PUWER Register's own listed serials against what the
            # individual audits found this week — catches contradictions either way
            # (PUWER lists a machine no individual audit confirms, or vice versa).
            # Skipped when PUWER Register only has a photo for this category (nothing
            # to read means nothing to reliably compare — would just be a false alarm).
            puwer_cross_check = None
            photo_uploaded = col_key in puwer_photo_only.get(gkey, set())
            if col_key in INDIVIDUAL_COVERABLE and not photo_uploaded:
                puwer_set = puwer_serials.get(gkey, {}).get(col_key, set())
                individual_set = machines_checked.get(gkey, {}).get(col_key, set())
                if puwer_set or individual_set:
                    in_puwer_not_individual = sorted(puwer_set - individual_set)
                    in_individual_not_puwer = sorted(individual_set - puwer_set)
                    if in_puwer_not_individual or in_individual_not_puwer:
                        puwer_cross_check = {
                            "in_puwer_not_individual": in_puwer_not_individual,
                            "in_individual_not_puwer": in_individual_not_puwer,
                        }

            if info and info["last_completed"]:
                if info.get("na"):
                    # PUWER Register confirms this site has none of this equipment —
                    # doesn't go stale the way a missed weekly inspection does.
                    status = "n/a"
                else:
                    try:
                        dc = datetime.fromisoformat(info["last_completed"].replace("Z", "+00:00")).date()
                    except Exception:
                        dc = None

                    if dc and dc >= week_start:
                        status = "ok"
                    elif dc:
                        status = "overdue"
                    else:
                        status = "missing"

                register_only = (
                    col_key in INDIVIDUAL_COVERABLE
                    and col_key in puwer_register_present.get(gkey, set())
                    and col_key not in direct_audit_cols.get(gkey, set())
                )
                # Title-parsed machine serials are far more reliable than item-text/notes
                # (which are often left blank) for the individual machine-type templates,
                # so merge them into found_serials — the same field the plant-register
                # Notes-column gap check (computeNotes, frontend) already compares.
                found = site_serials.get(gkey, {}).get(col_key, set()) | machines_checked.get(gkey, {}).get(col_key, set())
                templates_status[col_key] = {
                    "status": status,
                    "last_completed": info["last_completed"],
                    "audit_id": info["audit_id"],
                    "inspector": info["inspector"],
                    "found_serials": sorted(found),
                    "register_only": register_only,
                    "machines_checked": sorted(machines_checked.get(gkey, {}).get(col_key, set())),
                    "individual_audits": sorted(
                        individual_audits.get(gkey, {}).get(col_key, []),
                        key=lambda a: a["date_completed"] or "", reverse=True,
                    ),
                    "puwer_cross_check": puwer_cross_check,
                    "puwer_photo_uploaded": photo_uploaded,
                    "loler_report_uploaded": col_key == "LOLER" and gkey in loler_report_uploaded,
                }
            else:
                templates_status[col_key] = {
                    "status": "missing",
                    "last_completed": None,
                    "audit_id": None,
                    "inspector": None,
                    "found_serials": [],
                    "register_only": False,
                    "machines_checked": sorted(machines_checked.get(gkey, {}).get(col_key, set())),
                    "individual_audits": sorted(
                        individual_audits.get(gkey, {}).get(col_key, []),
                        key=lambda a: a["date_completed"] or "", reverse=True,
                    ),
                    "puwer_cross_check": puwer_cross_check,
                    "puwer_photo_uploaded": photo_uploaded,
                    "loler_report_uploaded": col_key == "LOLER" and gkey in loler_report_uploaded,
                }
        sites_list.append({"name": site_name, "job_code": job_code, "templates": templates_status})

    total     = len(sites_list)
    compliant = sum(1 for s in sites_list if all(t["status"] in ("ok", "n/a") for t in s["templates"].values()))

    result = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "week_start": from_date,
        "week_end": to_date,
        "sites": sites_list,
        "summary": {"total": total, "compliant": compliant, "has_gaps": total - compliant},
    }
    if search_errors:
        result["warnings"] = [f"{e['template']} search failed: {e['error']}" for e in search_errors]
    _write_cache(cache_key, result)
    return result


def parse_weekly_excel(excel_bytes):
    """Sites with an 'N' in any COL_COMPLIANCE column are required that week."""
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    _, _, header_row = build_lookup(ws)

    sites_data = []
    for row in ws.iter_rows(min_row=header_row + 1):
        site_name = str(row[0].value).strip() if row[0].value else ""
        if not site_name:
            continue
        job_code   = str(row[1].value).strip() if row[1].value else ""
        supervisor = str(row[11].value).strip() if row[11].value else ""

        required = [name for name, idx in COL_COMPLIANCE.items()
                    if row[idx].value and str(row[idx].value).strip().upper() == "N"]
        if required:
            sites_data.append({
                "site_name": site_name,
                "job_code": job_code,
                "supervisor": supervisor,
                "required_inspections": required,
            })
    return sites_data


# ── API Routes ────────────────────────────────────────────

@app.route("/api/health")
def api_health():
    """Liveness + config check — catches a missing SAFETYCULTURE_API_TOKEN
    (the exact cause of the earlier Vercel outage) without a live API call."""
    token_present = bool(os.environ.get("SAFETYCULTURE_API_TOKEN"))
    body = {"status": "ok" if token_present else "degraded", "safetyculture_token_configured": token_present}
    return jsonify(body), (200 if token_present else 503)


@app.route("/api/sites")
def api_sites():
    try:
        data = sc.get_sites()
        return jsonify(data)
    except Exception as e:
        app.logger.exception("api_sites failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/dashboard")
def api_dashboard():
    from_date = request.args.get("from")
    to_date   = request.args.get("to")
    if not from_date or not to_date:
        from_date, to_date = current_week_range()
    elif not valid_date_format(from_date, to_date):
        return jsonify({"error": "Dates must be in YYYY-MM-DD format"}), 400
    try:
        return jsonify(build_dashboard_data(from_date, to_date))
    except Exception as e:
        app.logger.exception("api_dashboard failed")
        return jsonify({"error": str(e)}), 500


_last_sync_at = 0.0
_SYNC_MIN_INTERVAL = 10  # seconds — bypasses the cache and fans out to 9 templates, don't let it be hammered

@app.route("/api/sync", methods=["POST"])
def api_sync():
    global _last_sync_at
    now = datetime.now(timezone.utc).timestamp()
    if now - _last_sync_at < _SYNC_MIN_INTERVAL:
        return jsonify({"error": f"Please wait a few seconds between syncs (max 1 per {_SYNC_MIN_INTERVAL}s)"}), 429
    _last_sync_at = now

    data = request.get_json(silent=True) or {}
    from_date = data.get("from")
    to_date   = data.get("to")
    if not from_date or not to_date:
        from_date, to_date = current_week_range()
    elif not valid_date_format(from_date, to_date):
        return jsonify({"error": "Dates must be in YYYY-MM-DD format"}), 400
    try:
        return jsonify(build_dashboard_data(from_date, to_date, force_refresh=True))
    except Exception as e:
        app.logger.exception("api_sync failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/validate-weekly", methods=["POST"])
def validate_weekly():
    """Cross-check an uploaded weekly Excel (sites marked 'N' = required) against SafetyCulture completions."""
    excel_file = request.files.get("excel")
    if not has_file(excel_file):
        return jsonify({"error": "No Excel file uploaded"}), 400

    from_date = request.form.get("from_date")
    to_date   = request.form.get("to_date")
    if not from_date or not to_date:
        from_date, to_date = current_week_range()
    elif not valid_date_format(from_date, to_date):
        return jsonify({"error": "Dates must be in YYYY-MM-DD format"}), 400

    try:
        sites_data = parse_weekly_excel(excel_file.read())
        if not sites_data:
            return jsonify({"error": "No sites found in Excel or no inspections marked with 'N'"}), 400

        dashboard = build_dashboard_data(from_date, to_date)
        by_job  = {s["job_code"]: s for s in dashboard["sites"] if s["job_code"]}
        by_name = {s["name"].lower(): s for s in dashboard["sites"]}

        results = []
        for site in sites_data:
            job     = normalize_job(site["job_code"]) if site["job_code"] else ""
            sc_site = by_job.get(job) or by_name.get(site["site_name"].lower())

            inspections = {}
            missing = []
            for insp in site["required_inspections"]:
                info = sc_site["templates"].get(insp) if sc_site else None
                done = bool(info) and info["status"] != "missing"
                inspections[insp] = {
                    "status": "done" if done else "missing",
                    "date_completed": info["last_completed"] if done else None,
                    "inspector": info["inspector"] if done else None,
                }
                if not done:
                    missing.append(insp)

            results.append({
                "site": site["site_name"],
                "job_code": site["job_code"],
                "supervisor": site["supervisor"],
                "inspections": inspections,
                "missing_inspections": missing,
                "notes": f"Missing: {', '.join(missing)}" if missing else "All inspections completed",
            })

        compliant = sum(1 for s in results if not s["missing_inspections"])
        return jsonify({
            "week_start": from_date,
            "week_end": to_date,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "total_sites": len(results),
            "sites": results,
            "summary": {"total": len(results), "compliant": compliant, "has_gaps": len(results) - compliant},
        })
    except Exception as e:
        app.logger.exception("validate_weekly failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/loler-check", methods=["POST"])
def loler_check():
    """Cross-check completed LOLER audits against the plant register PDF, item by item."""
    pdf_file = request.files.get("pdf")
    if not has_file(pdf_file):
        return jsonify({"error": "No PDF uploaded"}), 400

    from_date = request.form.get("from_date")
    to_date   = request.form.get("to_date")
    if not from_date or not to_date:
        from_date, to_date = current_week_range()
    elif not valid_date_format(from_date, to_date):
        return jsonify({"error": "Dates must be in YYYY-MM-DD format"}), 400

    try:
        sites = check_loler_coverage(from_date, to_date, pdf_file.read())
        return jsonify({
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "week_start": from_date,
            "week_end": to_date,
            "sites": sites,
        })
    except Exception as e:
        app.logger.exception("loler_check failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/inspections")
def api_inspections():
    site_filter     = request.args.get("site", "").lower()
    template_filter = request.args.get("template", "")
    from_date       = request.args.get("from")
    to_date         = request.args.get("to")

    if not from_date or not to_date:
        from_date, to_date = current_week_range()

    try:
        modified_after  = to_iso(from_date)
        modified_before = to_iso(to_date, end=True)

        templates = {tid: tkey for tid, tkey in TEMPLATES.items() if not template_filter or tkey == template_filter}
        audit_to_tkey, _ = search_templates(templates, modified_after, modified_before)

        details = sc.fetch_audits_parallel(list(audit_to_tkey.keys()))
        results = []
        for audit_id, detail in details.items():
            tkey = audit_to_tkey[audit_id]
            ad   = detail.get("audit_data", {})
            site_name = ad.get("site", {}).get("name", "").strip()
            if site_filter and site_filter not in site_name.lower():
                continue
            authorship = ad.get("authorship", {})
            results.append({
                "audit_id":       audit_id,
                "template":       tkey,
                "site":           site_name,
                "date_completed": ad.get("date_completed") or ad.get("date_modified", ""),
                "inspector":      authorship.get("author", "") if isinstance(authorship, dict) else "",
                "score":          ad.get("score", ""),
            })

        results.sort(key=lambda x: x["date_completed"] or "", reverse=True)
        return jsonify({"inspections": results})
    except Exception as e:
        app.logger.exception("api_inspections failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/inspections/<audit_id>")
def api_inspection_detail(audit_id):
    try:
        detail = sc.get_audit(audit_id)
        return jsonify(detail)
    except requests.HTTPError as e:
        status = e.response.status_code if e.response is not None else 502
        if status == 404:
            return jsonify({"error": "Inspection not found"}), 404
        if 400 <= status < 500:
            return jsonify({"error": str(e)}), status
        return jsonify({"error": str(e)}), 502
    except Exception as e:
        app.logger.exception("api_inspection_detail failed for %s", audit_id)
        return jsonify({"error": str(e)}), 500


# ── Plant register & active sites parsing ────────────────

@app.route("/api/parse-register", methods=["POST"])
def api_parse_register():
    f = request.files.get("pdf")
    if not has_file(f):
        return jsonify({"error": "No PDF uploaded"}), 400
    try:
        site_machines, name_to_job = parse_pdf(f.read())
        register = {
            job: {mtype: sorted(serials) for mtype, serials in mtypes.items()}
            for job, mtypes in site_machines.items()
        }
        return jsonify({"register": register, "name_to_job": name_to_job})
    except Exception as e:
        app.logger.exception("api_parse_register failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/parse-sites", methods=["POST"])
def api_parse_sites():
    f = request.files.get("excel")
    if not has_file(f):
        return jsonify({"error": "No Excel uploaded"}), 400
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()))
        ws = wb.active

        # Locate header row and column indexes
        header_row = 1
        site_col = job_col = sup_col = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                val = str(cell.value or "").strip().upper()
                if val == "SITE":
                    header_row = cell.row
                    site_col   = cell.column - 1
                elif val in ("JOB", "JOB CODE", "CODE"):
                    job_col = cell.column - 1
                elif "SUPER" in val or val in ("SSV", "SS", "SUPERVISOR"):
                    sup_col = cell.column - 1
            if site_col is not None:
                break

        if site_col is None:
            return jsonify({"error": "Could not find a 'Site' column in the first 10 rows of the Excel file"}), 400

        sites = []
        for row in ws.iter_rows(min_row=header_row + 1):
            def cv(idx):
                return str(row[idx].value or "").strip() if idx is not None and idx < len(row) else ""
            name = cv(site_col)
            if not name:
                continue
            job  = normalize_job(cv(job_col)) if job_col is not None else ""
            sup  = cv(sup_col) if sup_col is not None else ""
            sites.append({"name": name, "job_code": job, "supervisor": sup})

        return jsonify({"sites": sites})
    except Exception as e:
        app.logger.exception("api_parse_sites failed")
        return jsonify({"error": str(e)}), 500


# ── Excel report route (unchanged) ───────────────────────

@app.route("/generate", methods=["POST"])
def generate():
    from_date  = request.form.get("from_date")
    to_date    = request.form.get("to_date")
    excel_file = request.files.get("siteFile")
    pdf_file   = request.files.get("pdfFile")

    if not from_date or not to_date:
        return jsonify({"error": "Missing dates"}), 400
    if not valid_date_format(from_date, to_date):
        return jsonify({"error": "Dates must be in YYYY-MM-DD format"}), 400
    if not has_file(excel_file):
        return jsonify({"error": "Please upload your weekly Excel template"}), 400

    try:
        excel_bytes = excel_file.read()
        pdf_bytes   = pdf_file.read() if has_file(pdf_file) else None

        out = generate_report(from_date, to_date, excel_bytes, pdf_bytes)
        filename = f"Ashvale_Summary_{from_date}_to_{to_date}.xlsx"
        return send_file(out, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        app.logger.exception("generate failed")
        return jsonify({"error": str(e)}), 500


# ── Serve React app (catch-all for client-side routing) ──

@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_react(path):
    static_dir = app.static_folder
    if path and os.path.exists(os.path.join(static_dir, path)):
        return send_from_directory(static_dir, path)
    return send_from_directory(static_dir, "index.html")


if __name__ == "__main__":
    os.makedirs("static", exist_ok=True)
    print("Open: http://localhost:5050")
    app.run(port=5050, debug=False)
